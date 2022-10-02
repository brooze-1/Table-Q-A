# 开发时间：2022/8/6  9:10
# coding=utf-8
# [python实现——处理Excel表格（超详细）](https://blog.csdn.net/weixin_44288604/article/details/120731317)

import openpyxl

def num2alp():
    '''将阿拉伯数字映射到字母上'''
    dict_ ={}
    num = 1
    for i in range(65,91):
        dict_[num] = chr(i)
        num+=1

    # print(dict_)
    return dict_

def dict2file(sum_dict):
    '''将dict数据写入到文件中'''
    f = open("./res_dict/dict.txt","w",encoding="utf-8")
    for item in sum_dict:
        f.write(str(item))
        f.write("\n")
    f.close()


def process_sheet1(sheet1):
    '''处理第一张工作表 返回dict数据'''
    dict_ = num2alp()
    # temp_dict用于存储除字段外每行的信息
    temp_dict = {}
    # sum用于存储所有的temp_dict
    sum_dict = []
    # 获取表格的最大行数
    max_row = sheet1.max_row
    # print("max_row:",max_row)
    # 获取表格的最大列数
    max_column = sheet1.max_column
    # print("max_column:",max_column)
    for item in range(2, max_row+1):

        # 判断单元格内容是否存在 判断当前行表头是否存在
        if sheet1[f"A{item}"].value == None:
            continue

        # print(sheet1[f"A{item}"].value)
        for item_ in range(2,max_column+1):

            # 获取表头 例：张三
            table_header = sheet1[f"A{item}"].value
            # print("table_header:",table_header)

            # 获取字段 例：语文
            field = sheet1[f"{dict_[item_]}1"].value
            # print("field:",field)

            # 获取对应表头与字段的单元格的内容 例：98
            cell_content = sheet1[f"{dict_[item_]}{item}"].value
            # print("cell_content:",cell_content)

            # 例：{"张三语文":98}
            temp_dict[(str(table_header) + str(field))] = str(cell_content)

        # 将temp_dict存入sum列表中 例:temp_dict = {'张三语文': '98', '张三数学': '96', '张三英语': '88'}
    sum_dict.append(temp_dict)

    # print("sum_dict:\n",sum_dict)

    return sum_dict

def process_excel(excel_filepath,sheet_name="Sheet1"):
    '''对excel表格进行预处理,默认处理第一张sheet1表格'''
    workbook = openpyxl.load_workbook(excel_filepath)
    # print(workbook.sheetnames)  # 打印Excel表中的所有表
    # 处理第一张sheet表格
    sheet1 = workbook[f"{sheet_name}"]
    # 处理第一张工作表 返回dict数据
    sum_dict = process_sheet1(sheet1)
    # 将dict文件保存到文件中
    dict2file(sum_dict)


if __name__=="__main__":
    excel_filepath="./data/data.xlsx"
    # 对excel表格进行处理
    process_excel(excel_filepath)

