# 开发时间：2022/8/6  17:48
import openpyxl
import json
import ast

def read_question_dataset(question_filepath):
    '''读取待添加的问题和答案并转化为字典形式'''
    workbook = openpyxl.load_workbook(question_filepath)
    # print(workbook.sheetnames)  # 打印Excel表中的所有表
    sheet1 = workbook["Sheet1"]

    return sheet1


def read_dict_file(dict_filepath="./res_dict/dict.txt"):
    '''读取字典文件'''
    with open(dict_filepath, "r", encoding="utf-8") as f:
        content = f.read()
        if content == '':
            content = {"Hi":"Hello world !!!"}
            return content

        # content = json.loads(content)
        if content != '':
            content = ast.literal_eval(content)

        # ast.literal_eval('')
    return content


def QA2dict(sheet1):
    '''将问题和答案转换为字典的形式'''
    dict_ = {}
    max_row = sheet1.max_row
    for item in range(2, max_row+1):
        if sheet1[f"A{item}"].value == None:
            continue
        if sheet1[f"B{item}"].value == None:
            sheet1[f"B{item}"].value = "暂无相关数据"
        dict_[sheet1[f"A{item}"].value] = sheet1[f"B{item}"].value

    return dict_


def add_dict_to_file(dict_content,dict_filepath="./temp/dict3.txt"):
    '''将字典数据添加到问题库当中'''
    content = read_dict_file()
    # dict(list(defaults.items())+list(user.items()))
    # res = dict(list(content.items())+list(dict_content.items()))
    for k,v in content.items():
        dict_content.update({k:v})

    res = dict_content
    with open("./res_dict/dict.txt", "w", encoding="utf-8") as f:
        f.write(str(res))
    return None


def qu_process_sheet1(sheet1):
    '''处理第一张sheet表格'''
    dict_content = QA2dict(sheet1)
    add_dict_to_file(dict_content)


if __name__=="__main__":
    question_filepath = r"./data/data2.xlsx"
    sheet1 = read_question_dataset(question_filepath)
    qu_process_sheet1(sheet1)

